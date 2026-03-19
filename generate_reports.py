#!/usr/bin/env python3
"""
generate_reports.py
Legge FirewallPolicies_Full.json e genera:
  - FirewallPolicies_Report.html  (dashboard interattiva)
  - FirewallPolicies_Rules.csv    (flat CSV per Power BI / Excel)
  - FirewallPolicies_Report.xlsx  (Excel multi-foglio, richiede openpyxl)

Uso:
  python generate_reports.py <json_path> <output_folder>
"""

import json, sys, os, csv, datetime

# ── Input / Output ──────────────────────────────────────────────────────────
JSON_PATH   = sys.argv[1] if len(sys.argv) > 1 else "FirewallPolicies_Full.json"
OUTPUT_DIR  = sys.argv[2] if len(sys.argv) > 2 else os.path.dirname(os.path.abspath(JSON_PATH))
HTML_OUT    = os.path.join(OUTPUT_DIR, "FirewallPolicies_Report.html")
CSV_OUT     = os.path.join(OUTPUT_DIR, "FirewallPolicies_Rules.csv")
XLSX_OUT    = os.path.join(OUTPUT_DIR, "FirewallPolicies_Report.xlsx")

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Load JSON (handles UTF-8 BOM) ────────────────────────────────────────────
print(f"[1] Caricamento JSON: {JSON_PATH}")
with open(JSON_PATH, encoding='utf-8-sig') as f:
    raw = json.load(f)
policies = raw if isinstance(raw, list) else [raw]
print(f"    Policy caricate: {len(policies)}")

# ── Mappings ─────────────────────────────────────────────────────────────────
ACTION_MAP    = {"0":"Block","1":"Allow","2":"Allow (bypass)"}
DIRECTION_MAP = {"in":"Inbound","out":"Outbound","2":"Inbound","1":"Outbound"}
PROTO_MAP     = {"6":"TCP","17":"UDP","1":"ICMP","58":"ICMPv6","256":"Any","0":"Any"}

# ── Parser regole firewall ───────────────────────────────────────────────────
def extract_rules(settings):
    rules = []
    if settings is None:
        return rules
    # Normalizza: dict singolo -> lista
    if isinstance(settings, dict):
        settings = [settings]
    for s in settings:
        if not isinstance(s, dict):
            continue
        # Ogni elemento e {settingInstance: {...}, id: "N"}
        si = s.get('settingInstance', s)
        if not isinstance(si, dict):
            continue
        defid = si.get('settingDefinitionId', '')
        if 'firewallrules' not in defid.lower():
            continue
        coll = si.get('groupSettingCollectionValue', [])
        for item in coll:
            rule = {
                'Name':'', 'Protocol':'', 'LocalPorts':'', 'RemotePorts':'',
                'Direction':'', 'Action':'', 'Enabled':'SI',
                'AppPath':'', 'Description':'', 'RemoteAddresses':'', 'LocalAddresses':''
            }
            for child in item.get('children', []):
                cid  = child.get('settingDefinitionId', '').lower()
                sv   = child.get('simpleSettingValue', {})
                sval = sv.get('value', '') if isinstance(sv, dict) else ''
                cv   = child.get('choiceSettingValue', {})
                csuf = cv.get('value', '').split('_')[-1] if isinstance(cv, dict) and cv.get('value') else ''
                sscv = child.get('simpleSettingCollectionValue', [])
                if sscv:
                    sval = ', '.join(str(x.get('value','')) for x in sscv if isinstance(x, dict))

                if   cid.endswith('_name') and 'display' not in cid:
                    rule['Name'] = sval
                elif 'action_type' in cid:
                    rule['Action'] = ACTION_MAP.get(csuf, csuf)
                elif cid.endswith('_direction'):
                    rule['Direction'] = DIRECTION_MAP.get(csuf, csuf)
                elif cid.endswith('_protocol'):
                    pv = sval or csuf
                    rule['Protocol'] = PROTO_MAP.get(pv, pv)
                elif 'localport' in cid:
                    rule['LocalPorts'] = sval
                elif 'remoteport' in cid:
                    rule['RemotePorts'] = sval
                elif 'filepath' in cid or '_app_' in cid:
                    rule['AppPath'] = sval
                elif cid.endswith('_description'):
                    rule['Description'] = sval
                elif 'enabled' in cid:
                    rule['Enabled'] = 'NO' if (sval or csuf).lower() in ('false', '0') else 'SI'
                elif 'remoteaddress' in cid:
                    rule['RemoteAddresses'] = sval
                elif 'localaddress' in cid:
                    rule['LocalAddresses'] = sval

            if rule['Name']:
                rules.append(rule)
    return rules

# ── Arricchisci ogni policy ──────────────────────────────────────────────────
print("[2] Analisi regole e gruppi")
for p in policies:
    p['_rules']    = extract_rules(p.get('Settings'))
    p['_groups']   = p.get('AssignedGroups') or []
    if isinstance(p['_groups'], str):
        p['_groups'] = [p['_groups']] if p['_groups'] else []
    # Conta settings baseline (non-regole)
    settings = p.get('Settings')
    if isinstance(settings, dict): settings = [settings]
    baseline = 0
    for s in (settings or []):
        if isinstance(s, dict):
            si = s.get('settingInstance', s)
            if isinstance(si, dict) and 'firewallrules' not in si.get('settingDefinitionId','').lower():
                baseline += 1
    p['_baseline'] = baseline
    print(f"    {p.get('DisplayName','')} -> {len(p['_rules'])} regole, {p['_baseline']} settings baseline")

total_rules = sum(len(p['_rules']) for p in policies)
print(f"    Totale regole: {total_rules}")

# ── CSV ──────────────────────────────────────────────────────────────────────
print(f"[3] Generazione CSV: {CSV_OUT}")
HEADERS = ['PolicyName','PolicySource','PolicyTemplateId','LastModified','ScopeTags',
           'AssignedGroups','RuleName','Protocol','LocalPorts','RemotePorts',
           'Direction','Action','Enabled','RemoteAddresses','LocalAddresses','AppPath','Description']

def short_date(s):
    s = str(s or '')
    return s[:10] if len(s) >= 10 else s

with open(CSV_OUT, 'w', newline='', encoding='utf-8-sig') as f:
    w = csv.writer(f)
    w.writerow(HEADERS)
    for p in policies:
        base = [p.get('DisplayName',''), p.get('Source',''), p.get('TemplateId',''),
                short_date(p.get('LastModified','')), p.get('RoleScopeTagIds',''),
                ' | '.join(p['_groups'])]
        if p['_rules']:
            for r in p['_rules']:
                w.writerow(base + [r['Name'],r['Protocol'],r['LocalPorts'],r['RemotePorts'],
                                   r['Direction'],r['Action'],r['Enabled'],
                                   r['RemoteAddresses'],r['LocalAddresses'],r['AppPath'],r['Description']])
        else:
            w.writerow(base + ['(nessuna regola)','','','','','','','','','',''])
print(f"    OK ({sum(len(p['_rules']) or 1 for p in policies)} righe)")

# ── HTML ─────────────────────────────────────────────────────────────────────
print(f"[4] Generazione HTML: {HTML_OUT}")

def esc(s):
    return str(s or '').replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;')

gen_date   = datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
n_policies = len(policies)
n_sc       = sum(1 for p in policies if p.get('Source') == 'SettingsCatalog')
n_intent   = sum(1 for p in policies if p.get('Source') == 'Intent')

# Policy rows
pol_rows = ''
for i, p in enumerate(policies, 1):
    src = '<span class="bi">Intent</span>' if p.get('Source') == 'Intent' else '<span class="bsc">SC</span>'
    grp = ' | '.join(p['_groups']) if p['_groups'] else '(nessuno)'
    pol_rows += (f"<tr><td>{i}</td><td><b>{esc(p.get('DisplayName',''))}</b></td><td>{src}</td>"
                 f"<td class='m'>{esc(p.get('TemplateId',''))}</td>"
                 f"<td>{short_date(p.get('LastModified',''))}</td>"
                 f"<td>{esc(p.get('RoleScopeTagIds',''))}</td>"
                 f"<td class='c'>{p['_baseline']}</td>"
                 f"<td class='c'><b>{len(p['_rules'])}</b></td>"
                 f"<td>{esc(grp)}</td>"
                 f"<td class='d'>{esc(p.get('Description',''))}</td></tr>")

# Rule rows
rule_rows = ''
for p in policies:
    for r in p['_rules']:
        ab = '<span class="bbl">BLOCK</span>' if r['Action'] == 'Block' else '<span class="bal">ALLOW</span>' if r['Action'] else ''
        db = ('<span class="bin">IN</span>'  if r['Direction'] == 'Inbound'  else
              '<span class="bot">OUT</span>' if r['Direction'] == 'Outbound' else esc(r['Direction']))
        eb = '<span class="bdi">DISAB.</span>' if r['Enabled'] == 'NO' else '<span class="bok">ON</span>'
        op = ' style="opacity:.6"' if r['Enabled'] == 'NO' else ''
        rule_rows += (f"<tr{op}>"
                      f"<td><small>{esc(p.get('DisplayName',''))}</small></td>"
                      f"<td><b>{esc(r['Name'])}</b></td>"
                      f"<td class='c'>{esc(r['Protocol']) or '&mdash;'}</td>"
                      f"<td class='c'>{esc(r['LocalPorts']) or '&mdash;'}</td>"
                      f"<td class='c'>{esc(r['RemotePorts']) or '&mdash;'}</td>"
                      f"<td class='c'>{db}</td>"
                      f"<td class='c'>{ab}</td>"
                      f"<td class='c'>{eb}</td>"
                      f"<td class='d'>{esc(r['RemoteAddresses']) or '&mdash;'}</td>"
                      f"<td class='m d'>{esc(r['AppPath']) or '&mdash;'}</td>"
                      f"<td class='d'>{esc(r['Description']) or '&mdash;'}</td></tr>")

if not rule_rows:
    rule_rows = '<tr><td colspan="11" class="empty">Nessuna regola trovata nel JSON.</td></tr>'

# Group rows
grp_rows = ''
for p in policies:
    src = '<span class="bi">Intent</span>' if p.get('Source') == 'Intent' else '<span class="bsc">SC</span>'
    gh  = (' '.join(f"<span class='gtag'>{esc(g)}</span>" for g in p['_groups'])
           if p['_groups'] else '<span style="color:#ca5010">Nessun gruppo</span>')
    grp_rows += (f"<tr><td><b>{esc(p.get('DisplayName',''))}</b></td>"
                 f"<td>{src}</td><td>{gh}</td>"
                 f"<td>{esc(p.get('RoleScopeTagIds',''))}</td></tr>")

CSS = """*{box-sizing:border-box;margin:0;padding:0}
body{font-family:"Segoe UI",system-ui,sans-serif;background:#f3f2f1;color:#323130}
header{background:linear-gradient(135deg,#0078d4,#106ebe);color:#fff;padding:16px 28px;display:flex;align-items:center;gap:12px}
header h1{font-size:18px;font-weight:600}header small{font-size:11px;opacity:.8;display:block;margin-top:2px}.hico{font-size:30px}
.con{max-width:1600px;margin:0 auto;padding:16px}
.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:10px;margin-bottom:18px}
.kpi{background:#fff;border-radius:8px;padding:14px;box-shadow:0 1px 4px rgba(0,0,0,.1);text-align:center;border-top:4px solid #0078d4}
.kpi.g{border-top-color:#107c10}.kpi.o{border-top-color:#ca5010}
.kv{font-size:28px;font-weight:700;color:#0078d4}.kpi.g .kv{color:#107c10}.kpi.o .kv{color:#ca5010}
.kl{font-size:11px;color:#605e5c;margin-top:3px}
.tabs{display:flex;gap:3px;border-bottom:2px solid #0078d4}
.tab{padding:9px 18px;cursor:pointer;border-radius:6px 6px 0 0;font-size:13px;font-weight:600;color:#605e5c;background:#e1dfdd;border:1px solid #d1d1d1;border-bottom:none}
.tab.active{background:#0078d4;color:#fff;border-color:#0078d4}.tab:hover:not(.active){background:#c7e0f4;color:#0078d4}
.panel{display:none;background:#fff;border:1px solid #d1d1d1;border-top:none;border-radius:0 0 8px 8px;padding:16px;box-shadow:0 2px 8px rgba(0,0,0,.08)}.panel.active{display:block}
.tb{display:flex;flex-wrap:wrap;gap:8px;margin-bottom:12px;align-items:center}
.tb input,.tb select{padding:7px 10px;border:1px solid #d1d1d1;border-radius:4px;font-size:13px;min-width:160px}
.tb input:focus,.tb select:focus{outline:none;border-color:#0078d4}
.tw{overflow-x:auto;max-height:60vh;overflow-y:auto}
table{width:100%;border-collapse:collapse;font-size:12px}
th{background:#0078d4;color:#fff;padding:9px 10px;text-align:left;font-weight:600;position:sticky;top:0;white-space:nowrap;z-index:1}
td{padding:8px 10px;border-bottom:1px solid #edebe9;vertical-align:top}
tr:hover td{background:#f0f6fd}tr:nth-child(even) td{background:#faf9f8}tr:nth-child(even):hover td{background:#f0f6fd}
.bal{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:700;background:#dff6dd;color:#107c10}
.bbl{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:700;background:#fde7e9;color:#d83b01}
.bin{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:700;background:#c7e0f4;color:#0078d4}
.bot{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:700;background:#fff4ce;color:#8a4f00}
.bsc{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:700;background:#e3f2fd;color:#0078d4}
.bi{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:700;background:#fdecd8;color:#ca5010}
.bok{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:700;background:#dff6dd;color:#107c10}
.bdi{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:700;background:#fde7e9;color:#d83b01}
.c{text-align:center}.m{font-family:Consolas,monospace;font-size:11px;color:#605e5c;word-break:break-all}
.d{font-size:11px;color:#605e5c}.empty{text-align:center;color:#605e5c;padding:32px;font-style:italic}
.gtag{display:inline-block;background:#eff6fc;color:#0078d4;border:1px solid #c7e0f4;border-radius:10px;padding:2px 8px;font-size:11px;margin:2px}"""

JS = """function switchTab(n){
  var ns=["policies","rules","groups"];
  document.querySelectorAll(".tab").forEach(function(t,i){t.classList.toggle("active",ns[i]===n);});
  document.querySelectorAll(".panel").forEach(function(p){p.classList.remove("active");});
  document.getElementById("tab-"+n).classList.add("active");
}
function ft(id,q){
  q=q.toLowerCase();
  document.getElementById(id).querySelectorAll("tr").forEach(function(r){
    r.style.display=(!q||r.textContent.toLowerCase().includes(q))?"":"none";});
}
function filterRules(){
  var q=document.getElementById("rq").value.toLowerCase();
  var act=document.getElementById("ract").value.toLowerCase();
  var dir=document.getElementById("rdir").value.toLowerCase();
  var en=document.getElementById("ren").value;
  document.getElementById("rb").querySelectorAll("tr").forEach(function(r){
    var txt=r.textContent.toLowerCase();
    var show=(!q||txt.includes(q))&&(!act||txt.includes(act))&&(!dir||txt.includes(dir))&&(!en||r.innerHTML.includes(en));
    r.style.display=show?"":"none";});
}"""

html = f"""<!DOCTYPE html>
<html lang="it"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Intune Firewall Report</title>
<style>{CSS}</style></head><body>
<header><span class="hico">&#128737;</span><div>
<h1>Intune Endpoint Security &mdash; Firewall Policy Report</h1>
<small>Generato il {gen_date} &nbsp;|&nbsp; Python + PowerShell</small>
</div></header>
<div class="con">
<div class="kpis">
<div class="kpi"><div class="kv">{n_policies}</div><div class="kl">Policy totali</div></div>
<div class="kpi"><div class="kv">{n_sc}</div><div class="kl">Settings Catalog</div></div>
<div class="kpi o"><div class="kv">{n_intent}</div><div class="kl">Intent (legacy)</div></div>
<div class="kpi g"><div class="kv">{total_rules}</div><div class="kl">Regole firewall</div></div>
</div>
<div class="tabs">
<div class="tab active" onclick="switchTab('policies')">&#128203; Policy ({n_policies})</div>
<div class="tab" onclick="switchTab('rules')">&#128293; Regole ({total_rules})</div>
<div class="tab" onclick="switchTab('groups')">&#128101; Gruppi Assegnati</div>
</div>
<div id="tab-policies" class="panel active">
<div class="tb"><input type="text" id="pq" placeholder="&#128269; Cerca policy..." oninput="ft('pb',this.value)"></div>
<div class="tw"><table><thead><tr>
<th>#</th><th>Nome Policy</th><th>Sorgente</th><th>Template ID</th><th>Ultima Modifica</th>
<th>Scope Tags</th><th>Settings</th><th>Regole FW</th><th>Gruppi Assegnati</th><th>Descrizione</th>
</tr></thead><tbody id="pb">{pol_rows}</tbody></table></div></div>
<div id="tab-rules" class="panel">
<div class="tb">
<input type="text" id="rq" placeholder="&#128269; Cerca regola, porta, app..." oninput="filterRules()">
<select id="ract" onchange="filterRules()"><option value="">Tutte le azioni</option><option>Allow</option><option>Block</option></select>
<select id="rdir" onchange="filterRules()"><option value="">Tutte le direzioni</option><option>Inbound</option><option>Outbound</option></select>
<select id="ren" onchange="filterRules()"><option value="">Tutte</option><option value="bok">Solo abilitate</option><option value="bdi">Solo disabilitate</option></select>
</div>
<div class="tw"><table><thead><tr>
<th>Policy</th><th>Nome Regola</th><th>Prot.</th><th>Porte Locali</th><th>Porte Remote</th>
<th>Direzione</th><th>Azione</th><th>Abilitata</th><th>Indirizzi</th><th>App/Path</th><th>Descrizione</th>
</tr></thead><tbody id="rb">{rule_rows}</tbody></table></div></div>
<div id="tab-groups" class="panel">
<div class="tb"><input type="text" id="gq" placeholder="&#128269; Cerca gruppo o policy..." oninput="ft('gb',this.value)"></div>
<div class="tw"><table><thead><tr>
<th>Nome Policy</th><th>Sorgente</th><th>Gruppi Assegnati</th><th>Scope Tags</th>
</tr></thead><tbody id="gb">{grp_rows}</tbody></table></div></div>
</div>
<script>{JS}</script></body></html>"""

with open(HTML_OUT, 'w', encoding='utf-8') as f:
    f.write(html)
print(f"    OK")

# ── XLSX ─────────────────────────────────────────────────────────────────────
print(f"[5] Generazione Excel: {XLSX_OUT}")
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    BD = "0078D4"; BP = "C7E0F4"; GP = "DFF6DD"; RP = "FDE7E9"; OP = "FDECD8"; WH = "FFFFFF"

    def hdr(ws, row, col, text, bg=BD):
        c = ws.cell(row, col, text)
        c.font      = Font(name="Segoe UI", size=10, bold=True,
                           color="FFFFFF" if bg == BD else "000000")
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        s = Side(style="thin", color="D1D1D1")
        c.border    = Border(left=s, right=s, top=s, bottom=s)

    def cell(ws, row, col, val, bg=WH, bold=False):
        c = ws.cell(row, col, val)
        c.font      = Font(name="Segoe UI", size=10, bold=bold)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        s = Side(style="thin", color="D1D1D1")
        c.border    = Border(left=s, right=s, top=s, bottom=s)

    # Sheet 1: Policy
    ws1 = wb.active
    ws1.title = "Policy Riepilogo"
    ws1.freeze_panes = "A2"
    h1 = ["Nr","Nome Policy","Sorgente","Template ID","Ultima Modifica","Scope Tags","Settings","Regole FW","Gruppi","Descrizione"]
    w1 = [5,42,18,36,16,20,10,10,10,44]
    for i,(h,w) in enumerate(zip(h1,w1),1):
        hdr(ws1,1,i,h)
        ws1.column_dimensions[ws1.cell(1,i).column_letter].width = w
    ws1.row_dimensions[1].height = 22
    for r,p in enumerate(policies,2):
        bg = OP if p.get('Source')=='Intent' else (BP if r%2==0 else WH)
        cell(ws1,r,1,r-1,bg)
        cell(ws1,r,2,p.get('DisplayName',''),bg,True)
        cell(ws1,r,3,p.get('Source',''),bg)
        cell(ws1,r,4,p.get('TemplateId',''),bg)
        cell(ws1,r,5,short_date(p.get('LastModified','')),bg)
        cell(ws1,r,6,p.get('RoleScopeTagIds',''),bg)
        cell(ws1,r,7,p['_baseline'],bg)
        cell(ws1,r,8,len(p['_rules']),bg)
        cell(ws1,r,9,len(p['_groups']),bg)
        cell(ws1,r,10,p.get('Description',''),bg)
        ws1.row_dimensions[r].height = 18
    ws1.auto_filter.ref = f"A1:J{len(policies)+1}"

    # Sheet 2: Regole
    ws2 = wb.create_sheet("Regole Firewall")
    ws2.freeze_panes = "A2"
    h2 = ["Policy","Sorgente","Nome Regola","Protocollo","Porte Locali","Porte Remote","Direzione","Azione","Abilitata","Indirizzi","App/Path","Descrizione"]
    w2 = [35,16,35,12,16,16,12,12,10,30,36,36]
    for i,(h,w) in enumerate(zip(h2,w2),1):
        hdr(ws2,1,i,h)
        ws2.column_dimensions[ws2.cell(1,i).column_letter].width = w
    ws2.row_dimensions[1].height = 22
    r2 = 2
    for p in policies:
        for ru in p['_rules']:
            bg2 = (RP if ru['Action']=='Block' else
                   OP if ru['Enabled']=='NO'   else
                   BP if r2%2==0               else WH)
            cell(ws2,r2,1,p.get('DisplayName',''),bg2)
            cell(ws2,r2,2,p.get('Source',''),bg2)
            cell(ws2,r2,3,ru['Name'],bg2,True)
            cell(ws2,r2,4,ru['Protocol'],bg2)
            cell(ws2,r2,5,ru['LocalPorts'],bg2)
            cell(ws2,r2,6,ru['RemotePorts'],bg2)
            cell(ws2,r2,7,ru['Direction'],bg2)
            cell(ws2,r2,8,ru['Action'],bg2)
            cell(ws2,r2,9,ru['Enabled'],bg2)
            cell(ws2,r2,10,ru['RemoteAddresses'],bg2)
            cell(ws2,r2,11,ru['AppPath'],bg2)
            cell(ws2,r2,12,ru['Description'],bg2)
            ws2.row_dimensions[r2].height = 18
            r2 += 1
    if r2 > 2:
        ws2.auto_filter.ref = f"A1:L{r2-1}"

    # Sheet 3: Gruppi
    ws3 = wb.create_sheet("Gruppi Assegnati")
    ws3.freeze_panes = "A2"
    h3 = ["Nr","Nome Policy","Sorgente","Gruppo / Target","Scope Tags"]
    w3 = [5,44,18,42,24]
    for i,(h,w) in enumerate(zip(h3,w3),1):
        hdr(ws3,1,i,h)
        ws3.column_dimensions[ws3.cell(1,i).column_letter].width = w
    ws3.row_dimensions[1].height = 22
    r3 = 2; i3 = 0
    for p in policies:
        grps = p['_groups'] if p['_groups'] else ['(nessun gruppo assegnato)']
        for g in grps:
            i3 += 1
            bg3 = OP if g=='(nessun gruppo assegnato)' else (GP if r3%2==0 else WH)
            cell(ws3,r3,1,i3,bg3)
            cell(ws3,r3,2,p.get('DisplayName',''),bg3,True)
            cell(ws3,r3,3,p.get('Source',''),bg3)
            cell(ws3,r3,4,g,bg3)
            cell(ws3,r3,5,p.get('RoleScopeTagIds',''),bg3)
            ws3.row_dimensions[r3].height = 18
            r3 += 1
    if r3 > 2:
        ws3.auto_filter.ref = f"A1:E{r3-1}"

    wb.save(XLSX_OUT)
    print(f"    OK")

except ImportError:
    print("    [!!] openpyxl non installato - xlsx saltato")
    print("    Installa con: pip install openpyxl")
except Exception as e:
    print(f"    [!!] Errore Excel: {e}")

print(f"\nReport generati in: {OUTPUT_DIR}")
print(f"  FirewallPolicies_Report.html")
print(f"  FirewallPolicies_Rules.csv")
print(f"  FirewallPolicies_Report.xlsx")
